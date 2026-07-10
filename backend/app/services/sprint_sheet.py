"""Online sprint sheet — persisted, synced from DB, merges Asana changes automatically."""

from __future__ import annotations

import io
import re
import time
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.exc import OperationalError
from sqlalchemy.orm import Session, joinedload

from app.config import get_settings
from app.models import AsanaProject, JiraIssue, Module, SprintSheet, SprintSheetRow, Ticket
from app.services import google_sheets_sync as gsync
from app.services.section_utils import (
    display_pipeline_status,
    is_prioritized_section,
    is_sprint_pipeline_section,
    normalize_section_name,
    sprint_sheet_display_sort_key,
)
from app.services.ticket_parser import clean_title_for_release, extract_jira_key, jira_browse_url
from app.services.work_type import classify_work_type, work_type_bucket

settings = get_settings()


def _commit_with_retry(db: Session, retries: int = 4) -> None:
    for attempt in range(retries):
        try:
            db.commit()
            return
        except OperationalError as exc:
            if "locked" not in str(exc).lower() or attempt >= retries - 1:
                raise
            db.rollback()
            time.sleep(0.4 * (attempt + 1))

JIRA_URL_PATTERN = re.compile(
    r"https?://[^\s\)<>\]]*atlassian\.net/browse/[A-Z][A-Z0-9]+-\d+",
    re.I,
)

def _display_asana_priority(ticket: Ticket) -> str | None:
    """Asana Priority custom field (High / Medium / Low) — empty when not set."""
    raw = (ticket.asana_priority_raw or "").strip()
    return raw or None


SPRINT_COLUMNS = [
    ("title", "Ticket"),
    ("ticket_type", "Type"),
    ("priority", "Priority"),
    ("asana_link", "Asana Link"),
    ("doc_link", "Jira Link"),
    ("jira_status", "Jira Status"),
    ("dev_estimate", "Dev Est (hrs)"),
    ("qa_estimate", "QA Est (hrs)"),
    ("total_estimate", "Total Est (hrs)"),
    ("dev_assigned", "Dev Assigned"),
    ("qa_assigned", "QA Assigned"),
    ("status", "Status"),
]


def _extract_jira_link(title: str, description: str, jira_key: str | None) -> str | None:
    text = f"{title}\n{description or ''}"
    match = JIRA_URL_PATTERN.search(text)
    if match:
        return match.group(0).rstrip(".,;")
    if jira_key:
        return jira_browse_url(jira_key)
    key = extract_jira_key(title, description, settings.jira_project_key)
    return jira_browse_url(key)


def _jira_status_for_ticket(ticket: Ticket, jira_by_key: dict[str, JiraIssue]) -> str | None:
    if ticket.jira_issue:
        return ticket.jira_issue.status
    if ticket.jira_key:
        issue = jira_by_key.get(ticket.jira_key.upper())
        return issue.status if issue else None
    return None


def _ticket_base_row(
    ticket: Ticket,
    section: str,
    sheet_status: str,
    jira_by_key: dict[str, JiraIssue] | None = None,
) -> dict:
    title = clean_title_for_release(ticket.title, ticket.workshop_name)
    dev = ticket.dev_effort_hours
    qa = ticket.qa_effort_hours
    total = ticket.total_effort_hours
    if total is None and (dev is not None or qa is not None):
        total = (dev or 0) + (qa or 0)
    status_label = display_pipeline_status(section, sheet_status)
    jira_lookup = jira_by_key or {}
    work_type, type_label = classify_work_type(
        ticket.title,
        ticket.description,
        ticket.asana_type_raw,
        ticket.support_category,
    )
    return {
        "ticket_id": ticket.id,
        "asana_gid": ticket.asana_gid,
        "title": title,
        "ticket_type": type_label,
        "work_type": work_type,
        "dev_estimate": dev,
        "qa_estimate": qa,
        "total_estimate": total,
        "status": status_label,
        "doc_link": _extract_jira_link(title, ticket.description or "", ticket.jira_key),
        "jira_status": _jira_status_for_ticket(ticket, jira_lookup),
        "asana_link": ticket.asana_url,
        "dev_assigned": None,
        "qa_assigned": None,
        "sheet_status": sheet_status,
        "section_name": section,
        "asana_board_index": ticket.asana_board_index,
        "priority": _display_asana_priority(ticket),
    }


def _merge_row(base: dict, saved: dict | None) -> dict:
    if not saved:
        return base
    merged = {**base}
    manual_keys = (
        "status",
        "qa_estimate",
        "dev_assigned",
        "qa_assigned",
    )
    for key in manual_keys:
        val = saved.get(key)
        if val is not None and val != "":
            merged[key] = val
    dev = merged.get("dev_estimate")
    qa = merged.get("qa_estimate")
    if merged.get("total_estimate") is None and (dev is not None or qa is not None):
        try:
            merged["total_estimate"] = float(dev or 0) + float(qa or 0)
        except (TypeError, ValueError):
            pass
    return merged


def _hydrate_row_from_ticket(
    row: dict,
    ticket: Ticket | None,
    jira_by_key: dict[str, JiraIssue] | None = None,
) -> dict:
    """Always restore ticket title and type from Asana — never show type-only rows."""
    if not ticket:
        return row
    section = ticket.module.name if ticket.module else (row.get("section_name") or "Unknown")
    sheet_status = row.get("sheet_status") or "active"
    base = _ticket_base_row(ticket, section, sheet_status, jira_by_key)
    out = {**row}
    out["section_name"] = section
    out["title"] = base["title"] or ticket.title or "Untitled ticket"
    out["ticket_type"] = base["ticket_type"]
    out["work_type"] = base["work_type"]
    out["asana_link"] = base["asana_link"]
    out["doc_link"] = base["doc_link"]
    out["jira_status"] = base["jira_status"]
    out["dev_estimate"] = base["dev_estimate"] if out.get("dev_estimate") is None else out["dev_estimate"]
    if out.get("total_estimate") is None:
        out["total_estimate"] = base["total_estimate"]
    out["asana_board_index"] = ticket.asana_board_index
    out["priority"] = base["priority"]
    out["status"] = base["status"]
    return out


def _prioritized_board_rows(rows: list[dict]) -> list[dict]:
    """Active tickets in the Prioritized column, exact Asana board order."""
    active = [
        r for r in rows
        if r.get("sheet_status") != "removed" and is_prioritized_section(r.get("section_name"))
    ]
    active.sort(
        key=lambda r: (
            r.get("asana_board_index") if r.get("asana_board_index") is not None else 999999,
            r.get("ticket_id") or 0,
        )
    )
    return active


def _row_sort_key(row: dict) -> tuple:
    """Done and in-progress stages first; Prioritized (not started) last."""
    return sprint_sheet_display_sort_key(
        row.get("section_name"),
        row.get("asana_board_index"),
        row.get("ticket_id"),
    )


def _totals(rows: list[dict]) -> dict:
    active = [r for r in rows if r.get("sheet_status") != "removed"]
    dev = sum(float(r.get("dev_estimate") or 0) for r in active)
    qa = sum(float(r.get("qa_estimate") or 0) for r in active)
    total = sum(float(r.get("total_estimate") or 0) for r in active)

    def section_of(r: dict) -> str:
        return (r.get("section_name") or r.get("status") or "").strip()

    prioritized = sum(1 for r in active if is_prioritized_section(section_of(r)))
    done = sum(1 for r in active if normalize_section_name(section_of(r)) == normalize_section_name("Done"))
    in_progress = len(active) - prioritized - done

    prioritized_rows = [r for r in active if is_prioritized_section(section_of(r))]
    pri_bugs = [r for r in prioritized_rows if work_type_bucket(r) == "bug"]
    pri_reqs = [r for r in prioritized_rows if work_type_bucket(r) == "requirement"]

    def _hours(subset: list[dict], field: str = "total_estimate") -> float:
        return sum(float(r.get(field) or 0) for r in subset)

    return {
        "ticket_count": len(active),
        "prioritized": prioritized,
        "prioritized_bugs": len(pri_bugs),
        "prioritized_requirements": len(pri_reqs),
        "prioritized_other": len(prioritized_rows) - len(pri_bugs) - len(pri_reqs),
        "prioritized_bug_hours": _hours(pri_bugs),
        "prioritized_requirement_hours": _hours(pri_reqs),
        "prioritized_bug_dev_hours": _hours(pri_bugs, "dev_estimate"),
        "prioritized_requirement_dev_hours": _hours(pri_reqs, "dev_estimate"),
        "prioritized_bug_qa_hours": _hours(pri_bugs, "qa_estimate"),
        "prioritized_requirement_qa_hours": _hours(pri_reqs, "qa_estimate"),
        "in_progress": max(in_progress, 0),
        "done": done,
        "removed": sum(1 for r in rows if r.get("sheet_status") == "removed"),
        "dev_hours": dev,
        "qa_hours": qa,
        "total_hours": total,
        # legacy fields for older clients
        "in_sprint": prioritized,
        "released": 0,
    }


class SprintSheetService:
    def __init__(self, db: Session, project_gid: str | None = None):
        self.db = db
        self.project_gid = project_gid

    def _project(self) -> AsanaProject | None:
        if not self.project_gid:
            return None
        return self.db.query(AsanaProject).filter(AsanaProject.gid == self.project_gid).first()

    def _get_or_create_sheet(self, project_id: int, sprint_name: str) -> SprintSheet:
        sheet = (
            self.db.query(SprintSheet)
            .filter(SprintSheet.project_id == project_id, SprintSheet.name == sprint_name)
            .first()
        )
        if not sheet:
            sheet = SprintSheet(project_id=project_id, name=sprint_name, is_active=True)
            self.db.add(sheet)
            self.db.flush()
        sheet.updated_at = datetime.utcnow()
        return sheet

    def _ticket_section(self, ticket: Ticket) -> str:
        return ticket.module.name if ticket.module else "Unknown"

    def sync_sheet(self, sprint_name: str, section_name: str | None = None) -> SprintSheet:
        project = self._project()
        if not project:
            raise ValueError("Project not found")
        section = section_name or settings.asana_sprint_section_name
        sheet = self._get_or_create_sheet(project.id, sprint_name)

        saved_rows = {
            r.asana_gid: r
            for r in self.db.query(SprintSheetRow).filter(SprintSheetRow.sheet_id == sheet.id).all()
        }

        jira_by_key = {
            ji.jira_key.upper(): ji for ji in self.db.query(JiraIssue).all()
        }

        all_tickets = (
            self.db.query(Ticket)
            .options(joinedload(Ticket.module), joinedload(Ticket.jira_issue))
            .filter(
                Ticket.project_id == project.id,
                Ticket.asana_gid.isnot(None),
                Ticket.removed_from_asana.is_(False),
            )
            .all()
        )

        prioritized_gids = {
            t.asana_gid for t in all_tickets
            if t.module and is_sprint_pipeline_section(t.module.name)
        }

        for ticket in all_tickets:
            gid = ticket.asana_gid
            if not gid:
                continue
            section = self._ticket_section(ticket)
            saved = saved_rows.get(gid)

            if gid in prioritized_gids:
                sheet_status = "active"
            elif saved and saved.sheet_status in ("active", "in_sprint", "released"):
                sheet_status = "removed"
            elif saved:
                sheet_status = saved.sheet_status
            else:
                continue

            if saved is None and sheet_status == "removed":
                continue

            base = _ticket_base_row(ticket, section, sheet_status, jira_by_key)
            merged = _merge_row(base, saved.row_data if saved else None)

            if saved:
                saved.sheet_status = sheet_status
                saved.ticket_id = ticket.id
                saved.row_data = merged
                saved.updated_at = datetime.utcnow()
            else:
                self.db.add(SprintSheetRow(
                    sheet_id=sheet.id,
                    ticket_id=ticket.id,
                    asana_gid=gid,
                    sheet_status=sheet_status,
                    row_data=merged,
                ))

        for gid, saved in saved_rows.items():
            if gid in prioritized_gids:
                continue
            ticket_gone = saved.ticket and saved.ticket.removed_from_asana
            if saved.sheet_status in ("active", "in_sprint", "released") or ticket_gone:
                saved.sheet_status = "removed"
                section = self._ticket_section(saved.ticket) if saved.ticket else "Unknown"
                base = _ticket_base_row(saved.ticket, section, "removed", jira_by_key) if saved.ticket else saved.row_data
                saved.row_data = _merge_row(base, saved.row_data)
                saved.updated_at = datetime.utcnow()

        _commit_with_retry(self.db)
        return sheet

    def _sheet_is_linked(self, sheet: SprintSheet) -> bool:
        return bool(sheet.google_spreadsheet_id and gsync.is_configured())

    def _apply_pulled_rows(self, sheet: SprintSheet, pulled: list[dict]) -> bool:
        if not pulled:
            return False
        db_rows = self.db.query(SprintSheetRow).filter(SprintSheetRow.sheet_id == sheet.id).all()
        by_gid = {r.asana_gid: r for r in db_rows}
        changed = False
        for item in pulled:
            gid = item.get("asana_gid")
            if not gid or gid not in by_gid:
                continue
            db_row = by_gid[gid]
            merged = {**(db_row.row_data or {})}
            for key in ("status", "qa_estimate", "dev_assigned", "qa_assigned"):
                if key in item:
                    merged[key] = item[key]
            dev = merged.get("dev_estimate")
            qa = merged.get("qa_estimate")
            if merged.get("total_estimate") is None and (dev is not None or qa is not None):
                try:
                    merged["total_estimate"] = float(dev or 0) + float(qa or 0)
                except (TypeError, ValueError):
                    pass
            if merged != db_row.row_data:
                db_row.row_data = merged
                db_row.updated_at = datetime.utcnow()
                changed = True
        return changed

    def _apply_external_pull(self, sheet: SprintSheet) -> bool:
        if not sheet.google_spreadsheet_id or not gsync.is_configured():
            return False
        tab = sheet.google_tab_name or sheet.name
        try:
            pulled = gsync.pull_sheet(sheet.google_spreadsheet_id, tab)
            return self._apply_pulled_rows(sheet, pulled)
        except Exception:
            return False

    def _push_external(self, sheet: SprintSheet, payload: dict) -> None:
        if not sheet.google_spreadsheet_id or not gsync.is_configured():
            return
        tab = sheet.google_tab_name or sheet.name
        synced_at = gsync.push_sheet(sheet.google_spreadsheet_id, tab, payload)
        sheet.google_synced_at = synced_at
        sheet.google_sheet_url = gsync.sheet_url(sheet.google_spreadsheet_id, tab)
        sheet.updated_at = datetime.utcnow()

    def _apply_google_pull(self, sheet: SprintSheet) -> bool:
        return self._apply_external_pull(sheet)

    def _push_google(self, sheet: SprintSheet, payload: dict) -> None:
        self._push_external(sheet, payload)

    def _payload_from_sheet(self, sheet: SprintSheet, sprint_name: str, section: str, project: AsanaProject) -> dict:
        db_rows = (
            self.db.query(SprintSheetRow)
            .options(joinedload(SprintSheetRow.ticket).joinedload(Ticket.module))
            .filter(SprintSheetRow.sheet_id == sheet.id)
            .all()
        )
        order = {"active": 0, "in_sprint": 0, "removed": 1}
        db_rows.sort(
            key=lambda r: (
                order.get(r.sheet_status, 9),
                _row_sort_key({
                    **(r.row_data or {}),
                    "section_name": (
                        r.ticket.module.name
                        if r.ticket and r.ticket.module
                        else (r.row_data or {}).get("section_name")
                    ),
                    "asana_board_index": (
                        r.ticket.asana_board_index
                        if r.ticket
                        else (r.row_data or {}).get("asana_board_index")
                    ),
                    "ticket_id": r.ticket_id,
                }),
            )
        )
        rows = []
        jira_by_key = {ji.jira_key.upper(): ji for ji in self.db.query(JiraIssue).all()}
        for r in db_rows:
            row = dict(r.row_data or {})
            row["sheet_status"] = r.sheet_status
            row["sheet_row_id"] = r.id
            row = _hydrate_row_from_ticket(row, r.ticket, jira_by_key)
            rows.append(row)
        self._refresh_jira_on_rows(rows)
        rows.sort(key=_row_sort_key)
        req_rows = [r for r in rows if work_type_bucket(r) == "requirement"]
        bug_rows = [r for r in rows if work_type_bucket(r) == "bug"]
        prioritized_rows = _prioritized_board_rows(rows)
        return {
            "sheet_id": sheet.id,
            "sprint_name": sprint_name,
            "project_name": project.name,
            "project_gid": self.project_gid,
            "section": section,
            "generated_at": datetime.utcnow().isoformat(),
            "rows": rows,
            "prioritized_rows": prioritized_rows,
            "requirement_rows": req_rows,
            "bug_rows": bug_rows,
            "totals": _totals(rows),
            "asana_live": True,
            "persisted": True,
            "google_sheet_url": sheet.google_sheet_url,
            "google_synced_at": sheet.google_synced_at.isoformat() if sheet.google_synced_at else None,
            "google_sheets_configured": gsync.is_configured(),
            "google_service_account_email": gsync.service_account_email(),
            "sync_mode": "service_account" if sheet.google_spreadsheet_id else None,
        }

    def _refresh_jira_on_rows(self, rows: list[dict]) -> None:
        ticket_ids = [r["ticket_id"] for r in rows if r.get("ticket_id")]
        if not ticket_ids:
            return
        tickets = (
            self.db.query(Ticket)
            .options(joinedload(Ticket.module), joinedload(Ticket.jira_issue))
            .filter(Ticket.id.in_(ticket_ids))
            .all()
        )
        by_id = {t.id: t for t in tickets}
        jira_by_key = {ji.jira_key.upper(): ji for ji in self.db.query(JiraIssue).all()}
        for row in rows:
            ticket = by_id.get(row.get("ticket_id"))
            if not ticket:
                continue
            title = clean_title_for_release(ticket.title, ticket.workshop_name) or ticket.title or "Untitled ticket"
            work_type, type_label = classify_work_type(
                ticket.title,
                ticket.description,
                ticket.asana_type_raw,
                ticket.support_category,
            )
            row["title"] = title
            row["ticket_type"] = type_label
            row["work_type"] = work_type
            row["doc_link"] = _extract_jira_link(title, ticket.description or "", ticket.jira_key)
            row["jira_status"] = _jira_status_for_ticket(ticket, jira_by_key)

    def link_google_sheet(self, sprint_name: str, spreadsheet_url: str) -> dict:
        project = self._project()
        if not project:
            raise ValueError("Project not found")
        if not gsync.is_configured():
            raise ValueError(
                "Google Sheets is not configured on the server. "
                "Add GOOGLE_SERVICE_ACCOUNT_FILE to backend/.env and share your sheet with the service account email."
            )
        spreadsheet_id = gsync.parse_spreadsheet_id(spreadsheet_url)
        section = settings.asana_sprint_section_name
        sheet = self.sync_sheet(sprint_name, section)
        tab = gsync.sanitize_tab_name(sprint_name)
        sheet.google_spreadsheet_id = spreadsheet_id
        sheet.google_tab_name = tab
        sheet.google_sheet_url = spreadsheet_url.strip() or gsync.sheet_url(spreadsheet_id, tab)
        sheet.apps_script_url = None
        sheet.apps_script_secret = None
        self.db.commit()
        payload = self._payload_from_sheet(sheet, sprint_name, section, project)
        self._push_google(sheet, payload)
        self.db.commit()
        payload["google_sheet_url"] = sheet.google_sheet_url
        payload["google_synced_at"] = sheet.google_synced_at.isoformat() if sheet.google_synced_at else None
        payload["sync_mode"] = "service_account"
        return payload

    def build(self, sprint_name: str, section_name: str | None = None, *, refresh: bool = True) -> dict:
        project = self._project()
        if not project:
            return {"rows": [], "totals": _totals([])}
        section = section_name or settings.asana_sprint_section_name

        if refresh:
            sheet = self.sync_sheet(sprint_name, section)
            if self._sheet_is_linked(sheet):
                if self._apply_external_pull(sheet):
                    self.db.commit()
        else:
            sheet = self._get_or_create_sheet(project.id, sprint_name)
            has_rows = (
                self.db.query(SprintSheetRow)
                .filter(SprintSheetRow.sheet_id == sheet.id)
                .limit(1)
                .count()
                > 0
            )
            if not has_rows:
                sheet = self.sync_sheet(sprint_name, section)

        project = self._project()
        if not project:
            return {"rows": [], "totals": _totals([])}
        payload = self._payload_from_sheet(sheet, sprint_name, section, project)

        if refresh and self._sheet_is_linked(sheet):
            try:
                self._push_external(sheet, payload)
                self.db.commit()
                payload["google_sheet_url"] = sheet.google_sheet_url
                payload["google_synced_at"] = (
                    sheet.google_synced_at.isoformat() if sheet.google_synced_at else None
                )
                payload["sync_mode"] = "service_account"
            except Exception as exc:
                payload["google_sync_error"] = str(exc)
        elif self._sheet_is_linked(sheet):
            payload["google_sheet_url"] = sheet.google_sheet_url
            payload["google_synced_at"] = (
                sheet.google_synced_at.isoformat() if sheet.google_synced_at else None
            )
            payload["sync_mode"] = "service_account"

        payload["google_sheets_configured"] = gsync.is_configured()
        payload["google_service_account_email"] = gsync.service_account_email()
        return payload

    def save_rows(self, sprint_name: str, rows: list[dict]) -> dict:
        project = self._project()
        if not project:
            raise ValueError("Project not found")
        sheet = self._get_or_create_sheet(project.id, sprint_name)
        by_gid = {r.get("asana_gid"): r for r in rows if r.get("asana_gid")}

        db_rows = (
            self.db.query(SprintSheetRow)
            .options(joinedload(SprintSheetRow.ticket).joinedload(Ticket.module))
            .filter(SprintSheetRow.sheet_id == sheet.id)
            .all()
        )
        jira_by_key = {ji.jira_key.upper(): ji for ji in self.db.query(JiraIssue).all()}
        for db_row in db_rows:
            incoming = by_gid.get(db_row.asana_gid)
            if not incoming:
                continue
            merged = _merge_row(
                _hydrate_row_from_ticket(db_row.row_data or {}, db_row.ticket, jira_by_key),
                incoming,
            )
            if incoming.get("sheet_status"):
                db_row.sheet_status = incoming["sheet_status"]
            db_row.row_data = merged
            db_row.updated_at = datetime.utcnow()

        sheet.updated_at = datetime.utcnow()
        self.db.commit()
        result = self.build(sprint_name, refresh=False)
        if self._sheet_is_linked(sheet):
            try:
                self._push_external(sheet, result)
                self.db.commit()
                result["google_synced_at"] = (
                    sheet.google_synced_at.isoformat() if sheet.google_synced_at else None
                )
            except Exception as exc:
                result["google_sync_error"] = str(exc)
        return result

    def build_xlsx(self, payload: dict) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sprint Sheet"

        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14)
        meta_font = Font(size=10, color="4B5563")

        totals = payload["totals"]
        ws["A1"] = f"Sprint Sheet — {payload['sprint_name']}"
        ws["A1"].font = title_font
        ws.merge_cells("A1:J1")

        ws["A2"] = (
            f"Project: {payload.get('project_name') or '—'}  |  "
            f"Prioritized: {totals.get('prioritized', 0)}  |  "
            f"In progress: {totals.get('in_progress', 0)}  |  "
            f"Done: {totals.get('done', 0)}"
        )
        ws["A2"].font = meta_font
        ws.merge_cells("A2:J2")

        ws["A3"] = (
            f"Dev: {totals['dev_hours']:.0f} hrs  |  "
            f"QA: {totals['qa_hours']:.0f} hrs  |  "
            f"Total: {totals['total_hours']:.0f} hrs  |  "
            f"Status = current Asana board column"
        )
        ws["A3"].font = meta_font
        ws.merge_cells("A3:J3")

        header_row = 5
        headers = [label for _, label in SPRINT_COLUMNS]
        for col, label in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        keys = [key for key, _ in SPRINT_COLUMNS]
        row_num = 0
        for row in payload["rows"]:
            if row.get("sheet_status") == "removed":
                continue
            row_num += 1
            r = header_row + row_num
            display = {**row}
            for col, key in enumerate(keys, start=1):
                value = display.get(key)
                cell = ws.cell(row=r, column=col, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if key in ("dev_estimate", "qa_estimate", "total_estimate") and value is not None:
                    cell.number_format = "0.0"

        widths = [44, 12, 10, 36, 32, 14, 10, 10, 10, 14, 14, 18]
        for i, width in enumerate(widths[: len(headers)], start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()


def sync_all_sprint_sheets(db: Session, project_gid: str) -> list[dict]:
    """Rebuild every active in-app sprint sheet (and linked Google tabs) for this project."""
    project = db.query(AsanaProject).filter(AsanaProject.gid == project_gid).first()
    if not project:
        return []
    sheets = (
        db.query(SprintSheet)
        .filter(SprintSheet.project_id == project.id, SprintSheet.is_active.is_(True))
        .all()
    )
    if not sheets:
        return []
    svc = SprintSheetService(db, project_gid)
    results = []
    for sheet in sheets:
        try:
            svc.build(sprint_name=sheet.name, refresh=True)
            results.append({"sprint": sheet.name, "success": True})
        except Exception as exc:
            results.append({"sprint": sheet.name, "success": False, "error": str(exc)})
    return results


def sync_linked_google_sheets(db: Session, project_gid: str) -> list[dict]:
    """After Asana sync, refresh sprint sheets (includes Google Sheet push when linked)."""
    return sync_all_sprint_sheets(db, project_gid)
